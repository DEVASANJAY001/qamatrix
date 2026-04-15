import pathlib
from openpyxl import load_workbook
p = pathlib.Path('src/SQAM_Updatw.xlsm')
wb = load_workbook(p, data_only=True)
print('sheets', wb.sheetnames)
ws = wb[wb.sheetnames[0]]
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True)):
    print(i, row)
